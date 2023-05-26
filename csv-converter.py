import csv
from openpyxl import load_workbook

def extract_columns_from_excel(input_file, column_names):
    wb = load_workbook(filename=input_file)
    sheet = wb.active

    column_data = [[] for _ in range(len(column_names))]
    column_indices = [None] * len(column_names)

    for row in sheet.iter_rows(values_only=True:
        if not all(column_indices):
            for i, name in enumerate(column_names):
                if column_indices[i] is None:
                    try:
                        column_indices[i] = row.index(name)
                    except ValueError:
                        pass
        else:
            for i, col_index in enumerate(column_indices):
                if col_index < len(row):
                    value = row[col_index]
                    column_data[i].append(value)

    return column_data


input_file = 'input.xlsx' #enter your excel file name
column_names = ['Column1','Column2','Column3']#here you can add your column names

column_data = extract_columns_from_excel(input_file, column_names)

output_file = 'output.csv' #enter your csv output file name 
with open(output_file, 'w', newline='') as csv_file:
    writer = csv.writer(csv_file)
    writer.writerows(column_data)

